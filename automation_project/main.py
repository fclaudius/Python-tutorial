import openpyxl

excel_file = openpyxl.load_workbook("inventory.xlsx")
product_list = excel_file["Sheet1"]

products_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row,4).value
    inventory = product_list.cell(product_row,2).value
    price = product_list.cell(product_row,3).value
    product_num = product_list.cell(product_row,1).value
    inventory_price = product_list.cell(product_row,5)

    # calculation number of products per supplier
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] += + 1
    else:
        products_per_supplier[supplier_name] = 1

    # calculation total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # logic products with invenory less than 10
    if inventory < 10:
        products_under_10_inv[int(product_num)] = int(inventory)

    # add value for total price
    inventory_price.value = inventory * price

print(f"Products per Supplier: {products_per_supplier}")
print(f"Total value per supplier: {total_value_per_supplier}")
print(f"Products under 10: {products_under_10_inv}")

product_list.cell(1,5).value = "Calculated Price for each product"
excel_file.save("inventory_with_total_value.xlsx")

