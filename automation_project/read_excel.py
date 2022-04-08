import openpyxl

excel_file = openpyxl.load_workbook("inventory_with_total_value.xlsx")
product_list = excel_file["Sheet1"]

for product_row in range(1, product_list.max_row + 1):
    lineending = ";"
    for cell in range(1,5+1):
        print (f"cell={cell}",end=" ")
        if cell == 5:
            lineending = "\n"
        print (product_list.cell( product_row,cell).value, end=f"{lineending}")